# -*- coding: utf-8 -*-
'''
    Excle的配置文件
'''
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


class ExcelConf:
    def __init__(self):
        # 定义单元格样式：Pass Failed
        self.border_type_pass = Side(border_style=None, color='C6EFCE'),
        self.border_type_failed = Side(border_style=None, color='FFC7CE'),
        self.passStyle = {
            'border_pass': Border(left=self.border_type_pass, right=self.border_type_pass, top=self.border_type_pass,
                                  bottom=self.border_type_pass),
            'font_pass': Font(size=11, bold=True, name='宋体', color="006100"),
            'fill_pass': PatternFill(patternType="solid", start_color="C6EFCE"),
            'align': Alignment(horizontal='center', vertical='center', wrap_text=True)
        }
        self.failedStyle = {
            'border_failed': Border(left=self.border_type_failed, right=self.border_type_failed,
                                    top=self.border_type_failed,
                                    bottom=self.border_type_failed),
            'font_failed': Font(size=11, bold=True, name='宋体', color="9C0006"),
            'fill_failed': PatternFill(patternType="solid", start_color="FFC7CE"),
            'align': Alignment(horizontal='center', vertical='center', wrap_text=True)
        }

    # Pass配置
    def pass_(self, cell, row, column):
        # 写入Pass
        cell(row=row, column=column).value = 'Pass'
        # 写入单元格样式
        cell(row=row, column=column).border = self.passStyle['border_pass']
        cell(row=row, column=column).font = self.passStyle['font_pass']
        cell(row=row, column=column).fill = self.passStyle['fill_pass']
        cell(row=row, column=column).alignment = self.passStyle['align']

        # Failed配置
    def failed_(self, cell, row, column):
        # 写入Failed
        cell(row=row, column=column).value = 'Failed'
        # 写入单元格样式
        cell(row=row, column=column).border = self.passStyle['border_failed']
        cell(row=row, column=column).font = self.passStyle['font_failed']
        cell(row=row, column=column).fill = self.passStyle['fill_failed']
        cell(row=row, column=column).alignment = self.passStyle['align']
